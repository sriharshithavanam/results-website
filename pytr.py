import pandas as pd

import requests as req
import certifi
import urllib3
from requests.exceptions import MissingSchema

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import webbrowser

from bs4 import BeautifulSoup

import os

valueTable_CSE = {'171':'MATHEMATICS-I', '173':'CHEMISTRY', '175':'PROGRAMMING FOR PROBLEM SOLVING', '185':'ENVIRONMENTAL SCIENCE', '178':'CHEMISTRY LAB', '180':'PROGRAMMING FOR PROBLEM SOLVING LAB', '182':'WORKSHOP LAB', '176':'ESSENCE OF INDIAN TRADITIONAL KNOWLEDGE', '271':'ENGLISH', '272':'MATHEMATICS-II', '275':'PHYSICS', '276':'BASIC ELECTRICAL ENGINEERING', '287A':'ENGINEERING GRAPHICS LAB', '281':'ENGLISH LAB', '285':'PHYSICS LAB', '286':'BASIC ELECTRICAL ENGINEERING LAB', '279':'INDIAN CONSTITUTION'}
valueTable_IT = {'171':'MATHEMATICS-I', '172':'PHYSICS', '174':'BASIC ELECTRICAL ENGINEERING', '187':'ENGINEERING GRAPHICS', '186':'INDIAN CONSTITUTION', '179':'BASIC ELECTRICAL ENGINEERING LAB', '177':'PHYSICS LAB', '272':'MATHEMATICS-II', '273A':'ENGINEERING CHEMISTRY', '274':'PROGRAMMING FOR PROBLEM SOLVING', '284':'WORKSHOP LAB', '278':'ESSENCE OF INDIAN TRADITIONAL KNOWLEDGE', '281':'ENGLISH LAB', '282A':'ENGINEERING CHEMISTRY LAB', '283':'PROGRAMMING FOR PROBLEM SOLVING LAB', '277':'ENVIRONMENTAL SCIENCE'}

def extract_data(x):
    data = {}
    with open(x, 'r') as html_file:
        html_content = html_file.read()
        soup = BeautifulSoup(html_content, 'html.parser')
        gfg = soup.find(lambda tag: tag.name == "font" and "Name" in tag.text)
        v = list(gfg.parent.next_siblings)
        name = v[1].text
        gfg = soup.find_all(lambda tag: tag.name == "font" and "Grade Secured" in tag.text)
        for g in gfg:
            v = list(g.parent.parent.next_siblings)
            we_need = []
            for i in v:
                if(i != '\n'):
                    k = i.contents
                    for tag in k:
                        if tag != "\n":
                            we_need.append(tag)
            for i in range(0, len(we_need), 5):
                data[we_need[i].text.rstrip().lstrip().replace("\xa0", "")] = [we_need[i+3].text.rstrip().lstrip().replace("\xa0", ""), we_need[i+4].text.rstrip().lstrip().replace("\xa0", "")]
        gfg = soup.find(lambda tag: tag.name == "font" and "Result With SGPA" in tag.text)
        end_results = list(gfg.parent.parent.parent.next_siblings)
        semester = {}
        for end in end_results:
            if end != '\n':
                s = end.text.replace('\n', '').replace('   ', ' ').strip().split("  ") #we have double space left
                semester[s[0]] = [s[1]]
    print(name)
    print(data)
    print(semester)
    return name, data, semester


def xlw(l, s, valueTable):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    c = 1
    worksheet.cell(row=1, column=c).value = 'SNo'
    c += 1
    worksheet.cell(row=1, column=c).value = 'RNo'
    c += 1
    worksheet.cell(row=1, column=c).value = 'Name'
    c += 1
    for i in valueTable:
        worksheet.cell(row=1, column=c).value = i
        c += 1
        worksheet.merge_cells(start_row=1, start_column=c - 1, end_row=1, end_column=c)
        c += 1
    for i in range(1,9):
        worksheet.cell(row=1, column=c).value = i
        c += 1
    r = 2
    for i in range(0, len(l), 4):
        c = 1
        worksheet.cell(row=r, column=c).value = r - 1 #serial number
        c += 1
        worksheet.cell(row=r, column=c).value = l[i] #roll number
        c += 1
        worksheet.cell(row=r, column=c).value = l[i + 1] #name of the student
        c += 1
        while worksheet.cell(row=1, column=c).value != 1:
            try:
                worksheet.cell(row=r, column=c).value = l[i + 2][worksheet.cell(row=1, column=c).value][0]
                c += 1
                worksheet.cell(row=r, column=c).value = l[i + 2][worksheet.cell(row=1, column=c-1).value][1]
                c += 1
            except KeyError:
                worksheet.cell(row=r, column=c).value = 'NA'
                c += 1
                worksheet.cell(row=r, column=c).value = 'NA'
                c += 1
        while worksheet.cell(row=1, column=c).value != None:
            try:
                k = str(worksheet.cell(row=1, column=c).value)
                worksheet.cell(row=r, column=c).value = l[i + 3][k][0]
            except KeyError:
                worksheet.cell(row=r, column=c).value = 'NA'
            c += 1
        r+=1
    for i in range(4,c-9,2):
        worksheet.cell(row=1, column=i).value = valueTable[worksheet.cell(row=1, column=i).value]
    workbook.save(s)

l = []

r1 = int(input()) #input 1st roll number
r2 = int(input()) #input 2nd roll number

for i in range(r1, r2+1): #code to get the results
        if i == 245621733041:
            continue
        payload = {'mbstatus': 'SEARCH', 'htno': i, 'Submit.x': 32, 'Submit.y': 6}
        s = req.Session()
        s.cert = "C:/Users/BSV/PycharmProjects/ou-results/www.osmania.ac.in.crt"
        try:
            resp = req.post("https://www.osmania.ac.in/res07/20221206.jsp", data=payload, allow_redirects=True, verify=False)
            f = open("info.html", "w")
            f.write(resp.text)
            f.close()
            try:
                cert_reqs = 'CERT_REQUIRED'
                lname, ldata, lcgpa = extract_data("info.html")
                l.append(str(i))
                l.append(lname)
                l.append(ldata)
                l.append(lcgpa)
            except AttributeError:
                print("Page not found for " + str(i))
                break
        except ConnectionError:
            print("Connection failed")
            break
        # link error
        except MissingSchema:
            print("Incorrect URL. Please enter complete url in the format http://{url}")
            break
        except TimeoutError:
            print("Connection timed out")
            break

df = pd.read_excel("score_sheet.xlsx", 'Sheet')
json = df.to_json()

f = open('results.json', 'w')
f.write(json)